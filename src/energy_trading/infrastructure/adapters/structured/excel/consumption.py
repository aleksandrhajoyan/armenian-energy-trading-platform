"""Consumption Excel adapter: structured ``.xlsx`` source reader.

Workbook acquisition, worksheet selection, schema interpretation, optional
explicit unit/timezone normalization, row validation, and batch time-series
structural validation stay inside infrastructure. The application receives
only ``StructuredIngestionResult``. Units, timezones, and interval cadence
are never inferred. Excel formulas are not calculated.
"""

from __future__ import annotations

import asyncio
from collections.abc import Callable, Iterable, Iterator, Sequence
from datetime import UTC, datetime
from pathlib import Path
from typing import Protocol, cast
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils.exceptions import InvalidFileException  # type: ignore[import-untyped]
from pydantic import ValidationError

from energy_trading.application.errors import DependencyUnavailableError
from energy_trading.application.ports.structured_ingestion import StructuredIngestionResult
from energy_trading.domain.models.ingestion import AdapterDiagnostic, DiagnosticSeverity, DLQRecord
from energy_trading.domain.models.observations import ConsumptionRecord
from energy_trading.infrastructure.adapters.structured.consumption_mapping import (
    apply_consumption_unit_safety,
    canonical_column_index,
    default_consumption_field_specs,
    validated_consumption_field_specs,
)
from energy_trading.infrastructure.adapters.structured.consumption_values import (
    UnrecognizedConsumptionSourceValue,
    consumption_timestamp_from_excel,
)
from energy_trading.infrastructure.adapters.structured.normalization import (
    NormalizationConfigurationError,
    PowerUnit,
    SourceValueNormalizationError,
    normalize_power_to_mw,
    normalize_timestamp_to_utc,
    resolve_source_timezone,
)
from energy_trading.infrastructure.adapters.structured.schema_mapping import (
    CanonicalFieldSpec,
    DeterministicFieldResolver,
    FieldResolutionMethod,
    FieldResolutionStatus,
    SchemaResolution,
)
from energy_trading.infrastructure.adapters.structured.time_series import (
    MISSING_INTERVAL_GAP_CODE,
    ConsumptionGap,
    ConsumptionRecordCandidate,
    ConsumptionSeriesIssue,
    IntervalGrid,
    validate_consumption_series,
)

_ADAPTER_NAME = "consumption_excel"
_WORKBOOK_FORMAT_ERRORS = (BadZipFile, InvalidFileException, ParseError)

_MSG_MISSING_REQUIRED = "Excel schema is missing a required canonical field."
_MSG_AMBIGUOUS = "Excel schema contains an ambiguous header mapping."
_MSG_COLLISION = "Excel schema maps multiple columns to the same canonical field."
_MSG_UNRESOLVED = "Excel schema contains an unresolved extra column."
_MSG_FUZZY = "Excel header was resolved with deterministic fuzzy matching."
_MSG_ROW_SHAPE = "Excel row is missing a required mapped column."
_MSG_ROW_VALIDATION = "Excel row could not be converted to ConsumptionRecord."
_MSG_WORKBOOK_INVALID = "Excel workbook could not be parsed."
_MSG_WORKSHEET_MISSING = "Configured Excel worksheet is unavailable."
_MSG_UNAVAILABLE = "Excel source is unavailable"


class _WorkbookView(Protocol):
    """Narrow openpyxl workbook surface used at the vendor boundary."""

    sheetnames: object
    active: object | None

    def __getitem__(self, key: str) -> object: ...

    def close(self) -> None: ...


class _WorksheetView(Protocol):
    """Narrow openpyxl worksheet surface used at the vendor boundary."""

    def iter_rows(self, *args: object, **kwargs: object) -> Iterable[Iterable[object]]: ...


class ConsumptionExcelAdapter:
    """Infrastructure adapter that structurally satisfies ``StructuredIngestionPort``.

    Source path, optional worksheet name, and field-mapping configuration are
    constructor-injected. They are not part of the application-facing
    ``ingest()`` signature.
    """

    def __init__(
        self,
        *,
        path: Path,
        source_name: str,
        sheet_name: str | None = None,
        field_specs: tuple[CanonicalFieldSpec, ...] | None = None,
        clock: Callable[[], datetime] | None = None,
        source_power_unit: PowerUnit = PowerUnit.MW,
        source_timezone: str | None = None,
        interval_grid: IntervalGrid | None = None,
    ) -> None:
        self._path = _require_xlsx_path(path)
        cleaned_source = source_name.strip()
        if not cleaned_source:
            msg = "source_name must be a non-empty string"
            raise ValueError(msg)
        self._sheet_name = _optional_sheet_name(sheet_name)
        if not isinstance(source_power_unit, PowerUnit):
            raise NormalizationConfigurationError("source_power_unit must be a PowerUnit")
        self._source_power_unit = source_power_unit
        self._source_timezone = resolve_source_timezone(source_timezone)
        self._interval_grid = _optional_interval_grid(interval_grid)
        specs = (
            default_consumption_field_specs(source_power_unit)
            if field_specs is None
            else field_specs
        )
        self._source_name = cleaned_source
        self._clock = clock if clock is not None else _utc_now
        self._resolver = DeterministicFieldResolver(validated_consumption_field_specs(specs))

    @property
    def source_name(self) -> str:
        return self._source_name

    async def ingest(self) -> StructuredIngestionResult[ConsumptionRecord]:
        return await asyncio.to_thread(self._ingest_sync)

    def _ingest_sync(self) -> StructuredIngestionResult[ConsumptionRecord]:
        try:
            return self._read_and_normalize()
        except OSError as exc:
            raise DependencyUnavailableError(_MSG_UNAVAILABLE) from exc

    def _read_and_normalize(self) -> StructuredIngestionResult[ConsumptionRecord]:
        try:
            workbook = cast(
                _WorkbookView,
                load_workbook(
                    filename=self._path,
                    read_only=True,
                    data_only=True,
                ),
            )
        except _WORKBOOK_FORMAT_ERRORS:
            diagnostic = _error("excel_workbook_invalid", _MSG_WORKBOOK_INVALID)
            return self._result((), (diagnostic,), (self._source_dlq((diagnostic,)),))

        try:
            worksheet = _select_worksheet(workbook, self._sheet_name)
            if worksheet is None:
                diagnostic = _error("excel_worksheet_missing", _MSG_WORKSHEET_MISSING)
                return self._result((), (diagnostic,), (self._schema_dlq((diagnostic,)),))
            return self._ingest_worksheet(worksheet)
        finally:
            workbook.close()

    def _ingest_worksheet(
        self,
        worksheet: object,
    ) -> StructuredIngestionResult[ConsumptionRecord]:
        candidates: list[ConsumptionRecordCandidate] = []
        diagnostics: list[AdapterDiagnostic] = []
        dlq_records: list[DLQRecord] = []
        column_index: dict[str, int] = {}
        header_seen = False
        logical_row = 0
        for raw_row in _iter_value_rows(cast(_WorksheetView, worksheet)):
            if _is_blank_row(raw_row):
                continue
            logical_row += 1
            if not header_seen:
                header_seen = True
                header = _header_source_fields(raw_row)
                schema = apply_consumption_unit_safety(
                    self._resolver.resolve_schema(header),
                    source_power_unit=self._source_power_unit,
                )
                schema_diagnostics, fatal_errors = _schema_diagnostics(schema)
                diagnostics.extend(schema_diagnostics)
                if fatal_errors:
                    dlq_records.append(self._schema_dlq(fatal_errors))
                    return self._result([], diagnostics, dlq_records)
                column_index = canonical_column_index(schema)
                continue
            self._ingest_data_row(
                row=raw_row,
                logical_row=logical_row,
                column_index=column_index,
                candidates=candidates,
                diagnostics=diagnostics,
                dlq_records=dlq_records,
            )
        return self._with_series_validation(candidates, diagnostics, dlq_records)

    def _ingest_data_row(
        self,
        *,
        row: tuple[object, ...],
        logical_row: int,
        column_index: dict[str, int],
        candidates: list[ConsumptionRecordCandidate],
        diagnostics: list[AdapterDiagnostic],
        dlq_records: list[DLQRecord],
    ) -> None:
        try:
            consumer_id = row[column_index["consumer_id"]]
            raw_timestamp = row[column_index["timestamp"]]
            raw_mw = row[column_index["value_mw"]]
        except IndexError:
            diagnostic = _error("excel_row_shape_invalid", _MSG_ROW_SHAPE)
            diagnostics.append(diagnostic)
            dlq_records.append(self._row_dlq(logical_row, (diagnostic,)))
            return
        try:
            payload = {
                "consumer_id": consumer_id,
                "timestamp": normalize_timestamp_to_utc(
                    consumption_timestamp_from_excel(raw_timestamp),
                    self._source_timezone,
                ),
                "value_mw": normalize_power_to_mw(raw_mw, self._source_power_unit),
            }
            candidates.append(
                ConsumptionRecordCandidate(
                    record=ConsumptionRecord.model_validate(payload),
                    source_position=logical_row,
                )
            )
        except (
            UnrecognizedConsumptionSourceValue,
            SourceValueNormalizationError,
            ValidationError,
        ):
            diagnostic = _error("excel_row_validation_failed", _MSG_ROW_VALIDATION)
            diagnostics.append(diagnostic)
            dlq_records.append(self._row_dlq(logical_row, (diagnostic,)))

    def _with_series_validation(
        self,
        candidates: list[ConsumptionRecordCandidate],
        diagnostics: list[AdapterDiagnostic],
        dlq_records: list[DLQRecord],
    ) -> StructuredIngestionResult[ConsumptionRecord]:
        series = validate_consumption_series(candidates, self._interval_grid)
        for issue in series.issues:
            diagnostic = _series_diagnostic(issue)
            diagnostics.append(diagnostic)
            dlq_records.append(self._row_dlq(issue.source_position, (diagnostic,)))
        for gap in series.gaps:
            diagnostics.append(_gap_diagnostic(gap))
        return self._result(
            [candidate.record for candidate in series.valid_candidates],
            diagnostics,
            dlq_records,
        )

    def _result(
        self,
        records: Sequence[ConsumptionRecord],
        diagnostics: Sequence[AdapterDiagnostic],
        dlq_records: Sequence[DLQRecord],
    ) -> StructuredIngestionResult[ConsumptionRecord]:
        return StructuredIngestionResult(
            source_name=self._source_name,
            records=tuple(records),
            diagnostics=tuple(diagnostics),
            dlq_records=tuple(dlq_records),
        )

    def _schema_dlq(self, diagnostics: tuple[AdapterDiagnostic, ...]) -> DLQRecord:
        return self._dlq(
            record_id=f"{self._source_name}:schema",
            payload_reference=f"excel://{self._source_name}/schema",
            diagnostics=diagnostics,
        )

    def _source_dlq(self, diagnostics: tuple[AdapterDiagnostic, ...]) -> DLQRecord:
        return self._dlq(
            record_id=f"{self._source_name}:source",
            payload_reference=f"excel://{self._source_name}/source",
            diagnostics=diagnostics,
        )

    def _row_dlq(self, logical_row: int, diagnostics: tuple[AdapterDiagnostic, ...]) -> DLQRecord:
        return self._dlq(
            record_id=f"{self._source_name}:row:{logical_row}",
            payload_reference=f"excel://{self._source_name}/row/{logical_row}",
            diagnostics=diagnostics,
        )

    def _dlq(
        self,
        *,
        record_id: str,
        payload_reference: str,
        diagnostics: tuple[AdapterDiagnostic, ...],
    ) -> DLQRecord:
        return DLQRecord(
            record_id=record_id,
            failed_at=self._clock(),
            source_name=self._source_name,
            adapter_name=_ADAPTER_NAME,
            diagnostics=diagnostics,
            payload_reference=payload_reference,
        )


def _utc_now() -> datetime:
    return datetime.now(UTC)


def _require_xlsx_path(value: object) -> Path:
    if not isinstance(value, Path):
        msg = "path must be a pathlib.Path"
        raise TypeError(msg)
    if value.suffix.lower() != ".xlsx":
        msg = "Excel adapter accepts .xlsx sources only"
        raise ValueError(msg)
    return value


def _optional_interval_grid(value: IntervalGrid | None) -> IntervalGrid | None:
    if value is None:
        return None
    if not isinstance(value, IntervalGrid):
        raise NormalizationConfigurationError("interval_grid must be an IntervalGrid")
    return value


def _optional_sheet_name(value: str | None) -> str | None:
    if value is None:
        return None
    cleaned = value.strip()
    if not cleaned:
        msg = "sheet_name must be a non-empty string when provided"
        raise ValueError(msg)
    return cleaned


def _select_worksheet(workbook: _WorkbookView, sheet_name: str | None) -> object | None:
    if sheet_name is None:
        return workbook.active
    names_raw = workbook.sheetnames
    if not isinstance(names_raw, (list, tuple)):
        return None
    names = tuple(str(name) for name in names_raw)
    if sheet_name not in names:
        return None
    return workbook[sheet_name]


def _iter_value_rows(worksheet: _WorksheetView) -> Iterator[tuple[object, ...]]:
    for raw_row in worksheet.iter_rows(values_only=True):
        row: tuple[object, ...] = tuple(raw_row)
        yield row


def _is_blank_cell(value: object) -> bool:
    if value is None:
        return True
    return isinstance(value, str) and not value.strip()


def _is_blank_row(row: Sequence[object]) -> bool:
    return all(_is_blank_cell(cell) for cell in row)


def _trim_trailing_blank(row: Sequence[object]) -> tuple[object, ...]:
    end = len(row)
    while end > 0 and _is_blank_cell(row[end - 1]):
        end -= 1
    return tuple(row[:end])


def _header_source_fields(row: Sequence[object]) -> tuple[str, ...]:
    """String headers participate in resolution; other cells stay unresolved."""

    return tuple(_header_cell_as_source_field(cell) for cell in _trim_trailing_blank(row))


def _header_cell_as_source_field(cell: object) -> str:
    if isinstance(cell, str):
        return cell
    return ""


def _schema_diagnostics(
    schema: SchemaResolution,
) -> tuple[list[AdapterDiagnostic], tuple[AdapterDiagnostic, ...]]:
    diagnostics: list[AdapterDiagnostic] = []
    fatal: list[AdapterDiagnostic] = []
    for canonical in schema.missing_required_fields:
        diagnostic = AdapterDiagnostic(
            code="excel_missing_required_field",
            message=_MSG_MISSING_REQUIRED,
            severity=DiagnosticSeverity.ERROR,
            field_name=canonical,
        )
        diagnostics.append(diagnostic)
        fatal.append(diagnostic)
    for resolution in schema.field_resolutions:
        if resolution.status is FieldResolutionStatus.AMBIGUOUS:
            diagnostic = _error("excel_schema_ambiguous", _MSG_AMBIGUOUS)
            diagnostics.append(diagnostic)
            fatal.append(diagnostic)
        elif resolution.status is FieldResolutionStatus.UNRESOLVED:
            diagnostics.append(
                AdapterDiagnostic(
                    code="excel_unresolved_column",
                    message=_MSG_UNRESOLVED,
                    severity=DiagnosticSeverity.WARNING,
                    field_name=None,
                )
            )
        elif (
            resolution.status is FieldResolutionStatus.RESOLVED
            and resolution.method is FieldResolutionMethod.FUZZY
            and resolution.canonical_field is not None
        ):
            diagnostics.append(
                AdapterDiagnostic(
                    code="excel_fuzzy_field_resolution",
                    message=_MSG_FUZZY,
                    severity=DiagnosticSeverity.WARNING,
                    field_name=resolution.canonical_field,
                )
            )
    for collision in schema.collisions:
        diagnostic = AdapterDiagnostic(
            code="excel_schema_collision",
            message=_MSG_COLLISION,
            severity=DiagnosticSeverity.ERROR,
            field_name=collision.canonical_field,
        )
        diagnostics.append(diagnostic)
        fatal.append(diagnostic)
    return diagnostics, tuple(fatal)


def _error(code: str, message: str) -> AdapterDiagnostic:
    return AdapterDiagnostic(
        code=code,
        message=message,
        severity=DiagnosticSeverity.ERROR,
        field_name=None,
    )


def _series_diagnostic(issue: ConsumptionSeriesIssue) -> AdapterDiagnostic:
    return AdapterDiagnostic(
        code=issue.code,
        message=issue.message,
        severity=DiagnosticSeverity.ERROR,
        field_name=issue.field_name,
    )


def _gap_diagnostic(gap: ConsumptionGap) -> AdapterDiagnostic:
    return AdapterDiagnostic(
        code=MISSING_INTERVAL_GAP_CODE,
        message=gap.diagnostic_message(),
        severity=DiagnosticSeverity.ERROR,
        field_name="timestamp",
    )
